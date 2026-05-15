import os
import json
import random
import smtplib
import csv
import base64
import pickle
import cv2
import numpy as np
import urllib.request
from pathlib import Path
from datetime import datetime, timedelta
from email.mime.text import MIMEText

from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.contrib.auth.hashers import make_password, check_password
from django.utils import timezone
from django.db import transaction
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
from django.core.exceptions import ValidationError
from django.db import IntegrityError
from django.conf import settings
from django.db.models import Q
import traceback

from deepface import DeepFace

from colleges.models import College, Department, Class
from notifications.models import EmotionLog, Notification, ClassSchedule

from students.models import Student
from teachers.models import Teacher, CourseAllocation

from .decorators import student_required, teacher_required
from attendance.recognition import cosine_distance

from google import genai
from collections import Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
os.makedirs(DATA_DIR, exist_ok=True)

EMBED_FILE = os.path.join(DATA_DIR, "embeddings.pkl")
IDS_FILE = os.path.join(DATA_DIR, "ids.pkl")

ATTENDANCE_DIR = os.path.join(settings.BASE_DIR, "attendance", "attendance_records")

os.makedirs(ATTENDANCE_DIR, exist_ok=True)

COL_NAMES = ["Subject(Class)", "Enrollment", "Name", "Date", "Time"]

face_cascade = cv2.CascadeClassifier(
    cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
)

embedding_store = {}
embedding_counter = {}
embedding_saved = set()
otp_storage = {}

OTP_EXPIRY_MINUTES = 5

def face_capture(request):
    student_id = request.session.get("face_student_id")
    if not student_id:
        return redirect("student_signup")
    return render(request, "other/face_capture.html", {"student_id": student_id})

@csrf_exempt
def process_frame(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request"}, status=405)

    # ── Only allow face registration for session-authenticated students ──
    if not request.session.get("face_student_id"):
        return JsonResponse({"error": "Unauthorized"}, status=401)

    try:
        data = json.loads(request.body)
        student_id = str(data.get("student_id"))
        image_data = data.get("image")

        if not student_id or not image_data:
            return JsonResponse({"error": "Invalid data"}, status=400)

        if student_id in embedding_saved:
            return JsonResponse({"count": 30, "done": True})

        if student_id not in embedding_store:
            embedding_store[student_id] = []
            embedding_counter[student_id] = 0

        img_bytes = base64.b64decode(image_data.split(",")[1])
        frame = cv2.imdecode(np.frombuffer(img_bytes, np.uint8), cv2.IMREAD_COLOR)

        if embedding_counter[student_id] < 30:
            try:
                result = DeepFace.represent(
                    frame,
                    model_name="Facenet",
                    detector_backend="opencv",
                    enforce_detection=False,
                )

                if len(result) > 0:
                    embedding = result[0]["embedding"]
                    embedding_store[student_id].append(embedding)
                    embedding_counter[student_id] += 1
            except Exception as e:
                print("Embedding error:", e)

        count = embedding_counter[student_id]

        if count >= 30:
            embeddings_np = np.array(embedding_store[student_id])

            if os.path.exists(EMBED_FILE):
                with open(EMBED_FILE, "rb") as f:
                    old_embeds = pickle.load(f)
                embeddings_np = np.vstack((old_embeds, embeddings_np))

            with open(EMBED_FILE, "wb") as f:
                pickle.dump(embeddings_np, f)

            if os.path.exists(IDS_FILE):
                with open(IDS_FILE, "rb") as f:
                    ids = pickle.load(f)
            else:
                ids = []

            ids.extend([student_id] * 30)

            with open(IDS_FILE, "wb") as f:
                pickle.dump(ids, f)

            Student.objects.filter(student_id=student_id).update(face_registered=True)

            embedding_saved.add(student_id)
            del embedding_store[student_id]
            del embedding_counter[student_id]
            request.session.pop("face_student_id", None)

            try:
                from attendance.recognition import reload_embeddings
                reload_embeddings()
            except Exception as re_err:
                print("⚠️ Recognition reload error:", re_err)

            return JsonResponse({"count": 30, "done": True})

        return JsonResponse({"count": count, "done": False})

    except Exception as e:
        print("FACE ERROR:", e)
        return JsonResponse({"error": "Face processing failed"}, status=500)

def load_embeddings():
    import pickle

    known_embeddings = []
    known_ids = []

    if os.path.exists(EMBED_FILE) and os.path.exists(IDS_FILE):
        try:
            with open(EMBED_FILE, "rb") as f:
                known_embeddings = pickle.load(f)

            with open(IDS_FILE, "rb") as f:
                known_ids = pickle.load(f)

            print("✅ Embeddings loaded:", len(known_ids))
        except Exception as e:
            print("❌ Embedding load error:", e)
    else:
        print("❌ No embedding files found")

    return known_embeddings, known_ids

def attendance_page(request):

    student_id = request.session.get("student_id")

    if not student_id:
        return redirect("student_login")

    student = Student.objects.get(id=student_id)

    subject = request.GET.get("subject")
    schedule = request.GET.get("schedule")

    class_name = str(student.student_class)

    context = {"subject": subject, "class_name": class_name, "schedule": schedule}

    return render(request, "other/attendance.html", context)

@csrf_exempt
def mark_attendance(request):

    if request.method != "POST":
        return JsonResponse({"error": "Invalid request"}, status=400)

    student_session_id = request.session.get("student_id")

    if not student_session_id:
        return JsonResponse({"status": "login_required", "message": "Login required"})

    image_data   = request.POST.get("image")
    subject      = request.POST.get("subject", "Unknown").strip()
    class_name   = request.POST.get("class", "").strip().upper()
    schedule_id  = request.POST.get("schedule_id", "").strip()

    if not image_data:
        return JsonResponse({"error": "No image received"}, status=400)

    try:
        image_bytes = base64.b64decode(image_data.split(",")[1])
        np_arr = np.frombuffer(image_bytes, np.uint8)
        frame = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)
    except Exception as e:
        return JsonResponse({"status": "image_error"})

    # ── Use averaged recognition engine (higher accuracy) ──
    from attendance.recognition import recognize, reload_embeddings
    predicted_id = recognize(frame)

    if not predicted_id:
        return JsonResponse({"status": "unknown"})

    try:
        predicted_id = str(predicted_id).strip()
        student = Student.objects.get(student_id=predicted_id)
        student_name = f"{student.first_name} {student.last_name}"
        enrollment_no = (student.enrollment_no or "").strip()

    except Student.DoesNotExist:
        return JsonResponse({"status": "unknown"})

    if student.id != student_session_id:

        return JsonResponse(
            {
                "status": "unauthorized",
                "message": "You cannot mark attendance for another student",
            }
        )

    subject_class = subject

    if class_name:
        subject_class = f"{subject}({class_name})"

    now      = datetime.now()
    date_str = now.strftime("%d-%m-%Y")
    time_str = now.strftime("%H:%M:%S")

    subject_dir = os.path.join(ATTENDANCE_DIR, subject_class)
    os.makedirs(subject_dir, exist_ok=True)

    if schedule_id:
        file_path = os.path.join(subject_dir, f"Attendance_{date_str}_S{schedule_id}.csv")
    else:
        file_path = os.path.join(subject_dir, f"Attendance_{date_str}.csv")

    file_exists = os.path.isfile(file_path)

    if file_exists:
        with open(file_path, "r") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row["Enrollment"].strip() == enrollment_no:
                    print("⚠️ Attendance already marked for this session")
                    return JsonResponse({"status": "already_marked", "name": student_name})

    with open(file_path, "a", newline="") as csvfile:

        writer = csv.writer(csvfile)

        if not file_exists:
            writer.writerow(COL_NAMES)

        writer.writerow(
            [subject_class, enrollment_no, student_name, date_str, time_str]
        )

    print("✅ Attendance Marked:", student_name)

    return JsonResponse(
        {
            "status": "success",
            "name": student_name,
            "enrollment": enrollment_no,
            "subject": subject_class,
            "time": time_str,
        }
    )

def home(request):
    return render(request, "index.html")

def live_class(request):
    return render(request, "other/live_room.html")

@csrf_exempt
def teacher_verify_otp(request):
    if request.method != "POST":
        return JsonResponse(
            {"success": False, "error": "Invalid request method"}, status=405
        )

    try:
        data = json.loads(request.body)
        email = data.get("email")
        otp_input = data.get("otp")

        session_email = request.session.get("otp_email")
        session_otp = request.session.get("otp_code")
        session_expiry = request.session.get("otp_expiry")

        if not all([session_email, session_otp, session_expiry]):
            return JsonResponse(
                {"success": False, "error": "OTP expired or not sent"}, status=400
            )

        if email != session_email:
            return JsonResponse(
                {"success": False, "error": "Email mismatch"}, status=400
            )

        if timezone.now() > datetime.fromisoformat(session_expiry):
            request.session.pop("otp_email", None)
            request.session.pop("otp_code", None)
            request.session.pop("otp_expiry", None)
            return JsonResponse({"success": False, "error": "OTP expired"}, status=400)

        if otp_input != session_otp:
            return JsonResponse({"success": False, "error": "Invalid OTP"}, status=400)

        request.session["teacher_email_verified"] = email
        request.session["teacher_otp_verified_at"] = timezone.now().isoformat()

        request.session.pop("otp_code", None)
        request.session.pop("otp_expiry", None)

        return JsonResponse({"success": True, "message": "OTP verified"})

    except Exception as e:
        print("Error verifying OTP:", e)
        return JsonResponse({"success": False, "error": str(e)}, status=500)

@csrf_protect
def teacher_signup(request):
    print("\n" + "=" * 50)
    print("TEACHER SIGNUP FUNCTION CALLED")
    print("=" * 50)

    if request.method == "POST":
        try:
            data = request.POST

            first_name = data.get("first_name", "").strip()
            last_name = data.get("last_name", "").strip()
            email = data.get("email", "").strip()
            teacher_id = data.get("teacher_id", "").strip()

            password = data.get("password", "")
            confirm_password = data.get("confirm_password", "")

            college_id = data.get("college_id")
            college_code = data.get("college_code")
            college_name = data.get("college_name")

            department_id = data.get("department_id")
            department_name = data.get("department_name")

            qualification = data.get("qualification", "").strip()
            experience = data.get("experience", "").strip()

            terms_accepted = data.get("terms")

            print("FIRST NAME:", first_name)
            print("LAST NAME:", last_name)
            print("EMAIL:", email)
            print("TEACHER ID:", teacher_id)
            print("COLLEGE ID:", college_id)
            print("DEPARTMENT ID:", department_id)

            if not all(
                [
                    first_name,
                    last_name,
                    email,
                    teacher_id,
                    password,
                    confirm_password,
                ]
            ):
                print("❌ Missing required fields")
                return JsonResponse({"error": "Missing required fields"}, status=400)

            if password != confirm_password:
                print("❌ Passwords do not match")
                return JsonResponse({"error": "Passwords do not match"}, status=400)

            if Teacher.objects.filter(email=email).exists():
                return JsonResponse({"error": "Email already registered"}, status=400)

            if Teacher.objects.filter(teacher_id=teacher_id).exists():
                return JsonResponse(
                    {"error": "Teacher ID already registered"}, status=400
                )

            if not terms_accepted:
                return JsonResponse({"error": "Terms not accepted"}, status=400)

            verified_email = request.session.get("teacher_email_verified")
            otp_verified_at = request.session.get("teacher_otp_verified_at")

            if not verified_email or verified_email != email:
                return JsonResponse(
                    {"error": "Email not verified. Please complete OTP verification."},
                    status=403,
                )

            if not otp_verified_at:
                return JsonResponse({"error": "OTP verification required"}, status=403)

            verified_time = datetime.fromisoformat(otp_verified_at)
            expiry_time = verified_time + timedelta(minutes=5)

            if timezone.now() > expiry_time:
                request.session.pop("teacher_email_verified", None)
                request.session.pop("teacher_otp_verified_at", None)
                return JsonResponse(
                    {"error": "OTP verification expired. Please verify again."},
                    status=403,
                )

            print("\n----- DATABASE CHECKS -----")

            college = None

            if college_id:
                college = College.objects.filter(id=college_id).first()
            elif college_code and college_name:
                existing_college = College.objects.filter(
                    college_name=college_name
                ).first()
                if existing_college and existing_college.college_code != college_code:
                    return JsonResponse(
                        {"error": "College name already exists with different code"},
                        status=400,
                    )
                college, created = College.objects.get_or_create(
                    college_code=college_code, defaults={"college_name": college_name}
                )

            department = None

            if department_id:
                department = Department.objects.filter(id=department_id).first()
            elif department_name and college:
                department, created = Department.objects.get_or_create(
                    college=college, department_name=department_name
                )
                print("Department created:", created)

            print("---------------------------\n")

            with transaction.atomic():
                username_part = email.split("@")[0]
                username = f"{username_part}_{int(timezone.now().timestamp())}"

                teacher = Teacher.objects.create(
                    username=username,
                    email=email,
                    password=make_password(password),
                    first_name=first_name,
                    last_name=last_name,
                    teacher_id=teacher_id,
                    college=college,
                    department=department,
                    qualification=qualification if qualification else None,
                    experience=experience if experience else None,
                    email_verified=True,
                    terms_accepted=True,
                    is_active=True,
                )

                print(f"✅ Teacher created with ID: {teacher.id}")

            request.session.pop("teacher_email_verified", None)
            request.session.pop("teacher_otp_verified_at", None)

            if department:
                print("FINAL DEPARTMENT ID:", department.id)
                print("FINAL DEPARTMENT NAME:", department.department_name)

            print("\n✅ TEACHER REGISTRATION SUCCESSFUL")
            print("=" * 50 + "\n")

            return JsonResponse(
                {
                    "success": True,
                    "teacher_id": teacher.teacher_id,
                    "redirect": "/teacher-login/",
                }
            )

        except Exception as e:
            print("\n❌ TEACHER SIGNUP ERROR:")
            print(str(e))
            import traceback

            traceback.print_exc()

            return JsonResponse({"error": str(e)}, status=500)

    return render(request, "teacher/teacher_signup.html")

@csrf_exempt
def teacher_login(request):
    print("➡️ teacher_login view called")

    if request.method == "POST":
        print("✅ POST request received")

        try:
            email = request.POST.get("email")
            password = request.POST.get("password")

            print("📩 Email received:", email)
            print("🔑 Password received:", "YES" if password else "NO")

            if not email or not password:
                print("❌ Email or password missing")
                return JsonResponse(
                    {"success": False, "error": "Email and password are required"},
                    status=400,
                )

            teacher = Teacher.objects.filter(email=email).first()

            if not teacher:
                print("❌ Teacher not found for email:", email)
                return JsonResponse(
                    {"success": False, "error": "Teacher not found"}, status=404
                )

            print("✅ Teacher found:", teacher.full_name())

            if not teacher.is_active:
                print("❌ Teacher account inactive:", email)
                return JsonResponse(
                    {"success": False, "error": "Account is inactive"}, status=403
                )

            if not check_password(password, teacher.password):
                print("❌ Invalid password for:", email)
                return JsonResponse(
                    {"success": False, "error": "Invalid email or password"}, status=401
                )

            request.session.flush()

            request.session["role"] = "teacher"
            request.session["teacher_id"] = teacher.id
            request.session["first_name"] = (
                teacher.full_name().split()[0] if teacher.full_name() else ""
            )
            request.session["last_name"] = (
                " ".join(teacher.full_name().split()[1:])
                if len(teacher.full_name().split()) > 1
                else ""
            )

            teacher.last_login = timezone.now()
            teacher.save(update_fields=["last_login"])

            print("✅ Teacher login successful:", email)
            return JsonResponse({"success": True})

        except Exception as e:
            print("🔥 TEACHER LOGIN ERROR:", str(e))
            return JsonResponse(
                {"success": False, "error": "Internal server error"}, status=500
            )

    return render(request, "teacher/teacher_login.html")

@teacher_required
def teacher_dashboard(request):
    if "teacher_id" not in request.session:
        return redirect("/teacher-login/")

    attendance_data = []
    error = None

    try:
        teacher = Teacher.objects.select_related("college", "department").get(
            id=request.session["teacher_id"]
        )
    except Teacher.DoesNotExist:
        return redirect("/teacher-login/")

    if request.method == "POST":
        subject = request.POST.get("subject")
        date_input = request.POST.get("date")

        print(f"\n{'='*50}")
        print(f"ATTENDANCE SEARCH:")
        print(f"Subject from form: '{subject}'")
        print(f"Date from form: '{date_input}'")
        print(f"{'='*50}\n")

        if subject and date_input:
            formatted_date = datetime.strptime(date_input, "%Y-%m-%d").strftime(
                "%d-%m-%Y"
            )
            print(f"Formatted date: '{formatted_date}'")

            allocation_check = CourseAllocation.objects.filter(
                teacher_id=request.session["teacher_id"], subject_name=subject
            )
            print(f"Allocation exists: {allocation_check.exists()}")

            if not allocation_check.exists():
                error = "You are not assigned to this subject."
            else:
                allocation = (
                    CourseAllocation.objects.filter(
                        teacher_id=request.session["teacher_id"], subject_name=subject
                    )
                    .select_related("class_assigned", "class_assigned__department")
                    .first()
                )

                if allocation:
                    class_obj = allocation.class_assigned
                    print(f"Found allocation - Class: '{class_obj.class_name}'")
                    print(f"Department: '{class_obj.department.department_name}'")

                    full_class_name = f"{class_obj.class_name} - {class_obj.department.department_name}"
                    subject_class = f"{subject}({full_class_name})"

                    print(f"Looking for subject_class: '{subject_class}'")

                    print(f"\nATTENDANCE_DIR: '{ATTENDANCE_DIR}'")
                    if os.path.exists(ATTENDANCE_DIR):
                        print(f"Contents of ATTENDANCE_DIR:")
                        for item in os.listdir(ATTENDANCE_DIR):
                            print(f"  - {item}")

                    parent_dir = os.path.join(ATTENDANCE_DIR, subject_class)

                    print(f"\nLooking in dir: '{parent_dir}'")
                    print(f"Dir exists: {os.path.exists(parent_dir)}")

                    if os.path.exists(parent_dir):
                        # Collect all CSV files for this date (handles _S1, _S2, ... suffixes)
                        matching_files = sorted([
                            f for f in os.listdir(parent_dir)
                            if f.startswith(f"Attendance_{formatted_date}") and f.endswith(".csv")
                        ])
                        print(f"Matching files: {matching_files}")

                        if matching_files:
                            seen_enrollments = set()
                            for fname in matching_files:
                                fpath = os.path.join(parent_dir, fname)
                                try:
                                    with open(fpath, newline="", encoding="utf-8") as file:
                                        reader = csv.DictReader(file)
                                        for row in reader:
                                            enrollment = row.get("Enrollment", "").strip()
                                            if enrollment and enrollment not in seen_enrollments:
                                                seen_enrollments.add(enrollment)
                                                attendance_data.append(
                                                    {
                                                        "enrollment": enrollment,
                                                        "name": row.get("Name", ""),
                                                        "time": row.get("Time", ""),
                                                    }
                                                )
                                except Exception as e:
                                    print(f"Error reading {fname}: {e}")

                            print(f"Total unique records found: {len(attendance_data)}")

                            if not attendance_data:
                                error = "No attendance records found for this date."
                            else:
                                print(f"Successfully loaded {len(attendance_data)} records")
                        else:
                            print(f"No files found for date {formatted_date}")
                            error = f"Attendance file not found for {subject} on {formatted_date}."
                    else:
                        print(f"\nDirectory does NOT exist: '{parent_dir}'")
                        error = f"Attendance file not found for {subject} on {formatted_date}."
                else:
                    error = "Could not determine class for this subject."

    context = {
        "attendance_data": attendance_data,
        "error": error,
        "teacher": teacher,
        "first_name": request.session.get("first_name"),
        "last_name": request.session.get("last_name"),
        "teacher_id": request.session.get("teacher_id"),
    }

    return render(request, "teacher/teacher_dashboard.html", context)

@csrf_protect
def teacher_reset_password(request):
    print("➡️ teacher_reset_password VIEW CALLED")

    if request.method == "POST":
        print("✅ POST request received")

        try:
            data = json.loads(request.body)
            print("📦 Request JSON:", data)

            action = data.get("action")
            email = data.get("email", "").strip()

            print("🔧 Action:", action)
            print("📧 Email:", email)

            if action == "send_otp":
                print("📨 STEP 1 : SEND OTP")

                if not email:
                    print("❌ Email missing")
                    return JsonResponse({"error": "Email required"}, status=400)

                teacher = Teacher.objects.filter(email=email).first()
                if not teacher:
                    print("❌ Email not registered:", email)
                    return JsonResponse({"error": "Email not registered"}, status=404)

                otp = str(random.randint(100000, 999999))
                expiry = timezone.now() + timedelta(minutes=5)

                request.session["reset_otp"] = otp
                request.session["reset_otp_email"] = email
                request.session["reset_otp_expiry"] = expiry.isoformat()
                request.session["reset_otp_role"] = "teacher"

                print("🔐 OTP Generated:", otp)
                print("⏰ OTP Expiry:", expiry)

                email_status = send_email_otp(
                    email, otp, purpose="forgot", role="teacher"
                )
                print("📤 Email send status:", email_status)

                if not email_status:
                    print("❌ Failed to send OTP email")
                    return JsonResponse({"error": "Failed to send OTP"}, status=500)

                return JsonResponse(
                    {"success": True, "message": "OTP sent successfully"}
                )

            elif action == "verify_otp":
                print("🔍 STEP 2 : VERIFY OTP")

                otp_input = data.get("otp", "").strip()

                session_otp = request.session.get("reset_otp")
                session_email = request.session.get("reset_otp_email")
                session_expiry = request.session.get("reset_otp_expiry")

                print("🔑 OTP Entered:", otp_input)
                print("📂 Session OTP:", session_otp)
                print("📧 Session Email:", session_email)
                print("⏳ Session Expiry:", session_expiry)

                if not (session_otp and session_email and session_expiry):
                    print("❌ OTP session missing")
                    return JsonResponse(
                        {"error": "OTP expired or not sent"}, status=403
                    )

                if email != session_email:
                    print("❌ Email mismatch")
                    return JsonResponse({"error": "Email mismatch"}, status=403)

                if timezone.now() > datetime.fromisoformat(session_expiry):
                    print("❌ OTP expired by time")
                    request.session.pop("reset_otp", None)
                    request.session.pop("reset_otp_email", None)
                    request.session.pop("reset_otp_expiry", None)
                    return JsonResponse({"error": "OTP expired"}, status=403)

                if otp_input != session_otp:
                    print("❌ Invalid OTP")
                    return JsonResponse({"error": "Invalid OTP"}, status=403)

                request.session["reset_email_verified"] = email
                request.session.pop("reset_otp", None)
                request.session.pop("reset_otp_expiry", None)

                print("✅ OTP verified successfully")

                return JsonResponse({"success": True, "message": "OTP verified"})

            elif action == "reset_password":
                print("🔁 STEP 3 : RESET PASSWORD")

                verified_email = request.session.get("reset_email_verified")
                print("📂 Verified email in session:", verified_email)

                if verified_email != email:
                    print("❌ OTP verification required")
                    return JsonResponse(
                        {"error": "OTP verification required"}, status=403
                    )

                new_password = data.get("password", "").strip()
                print("🔑 New password length:", len(new_password))

                if not new_password or len(new_password) < 8:
                    print("❌ Password validation failed")
                    return JsonResponse(
                        {"error": "Password must be at least 8 characters"}, status=400
                    )

                teacher = Teacher.objects.filter(email=email).first()
                if not teacher:
                    print("❌ Teacher not found")
                    return JsonResponse({"error": "Email not found"}, status=404)

                teacher.password = make_password(new_password)
                teacher.save(update_fields=["password"])

                request.session.pop("reset_email_verified", None)

                print("✅ Password reset successful for:", email)

                return JsonResponse(
                    {"success": True, "message": "Password reset successful"}
                )

            else:
                print("❌ Invalid action")
                return JsonResponse({"error": "Invalid action"}, status=400)

        except Exception as e:
            print("🔥 TEACHER RESET PASSWORD ERROR 🔥")
            print("Exception:", str(e))
            return JsonResponse({"error": "Internal server error"}, status=500)

    print("📄 Rendering teacher_reset_password.html")
    return render(request, "teacher/teacher_reset_password.html")

@csrf_exempt
def teacher_send_otp(request):
    if request.method != "POST":
        return JsonResponse(
            {"success": False, "error": "Invalid request method"}, status=405
        )

    try:
        data = json.loads(request.body)
        email = data.get("email")
        if not email:
            return JsonResponse(
                {"success": False, "error": "Email is required"}, status=400
            )

        if Teacher.objects.filter(email=email).exists():
            return JsonResponse(
                {"success": False, "error": "Email already registered"}, status=400
            )

        otp = str(random.randint(100000, 999999))
        expiry = timezone.now() + timedelta(minutes=5)

        request.session["otp_email"] = email
        request.session["otp_code"] = otp
        request.session["otp_expiry"] = expiry.isoformat()
        request.session["otp_purpose"] = "signup"
        request.session["otp_role"] = "teacher"

        mail_sent = send_email_otp(email, otp, purpose="signup", role="teacher")
        if not mail_sent:
            return JsonResponse({"error": "Failed to send OTP"}, status=500)

        return JsonResponse({"success": True})

    except Exception as e:
        print("Error sending teacher OTP:", e)
        return JsonResponse({"success": False, "error": str(e)}, status=500)

@csrf_exempt
def get_teacher_courses(request, teacher_id):
    session_teacher_id = request.session.get("teacher_id")
    if not session_teacher_id or int(session_teacher_id) != int(teacher_id):
        return JsonResponse({"success": False, "error": "Unauthorized"}, status=401)
    try:
        courses = CourseAllocation.objects.filter(
            teacher_id=teacher_id, is_active=True
        ).select_related("class_assigned")

        data = []
        for course in courses:
            data.append(
                {
                    "id": course.id,
                    "display_name": f"{course.class_assigned.class_name} - {course.subject_name}",
                    "class_name": course.class_assigned.class_name,
                    "subject": course.subject_name,
                }
            )

        return JsonResponse({"success": True, "courses": data})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})

@csrf_exempt
def add_schedule(request):
    if request.method == "POST":
        # ── Session auth check ──
        teacher_id = request.session.get("teacher_id")
        if not teacher_id:
            return JsonResponse({"success": False, "error": "Unauthorized"}, status=401)

        try:
            data = json.loads(request.body)
            course_id = data.get("course_id")
            date = data.get("date")
            start_time = data.get("start_time")

            if not all([teacher_id, course_id, date, start_time]):
                return JsonResponse(
                    {"success": False, "error": "All fields are required"}
                )

            course = CourseAllocation.objects.select_related(
                "class_assigned", "department", "college"
            ).get(id=course_id, teacher_id=teacher_id)

            schedule = ClassSchedule.objects.create(
                teacher_id=teacher_id,
                course_allocation=course,
                date=date,
                start_time=start_time,
            )

            class_obj = course.class_assigned
            students = Student.objects.filter(student_class=class_obj)

            notifications = []
            for student in students:
                notifications.append(
                    Notification(
                        sent_by_id=teacher_id,
                        student=student,
                        class_target=class_obj,
                        title=f"New Class Scheduled: {course.subject_name}",
                        message=f"Class for {class_obj.class_name} scheduled on {date} at {start_time}",
                        notification_type="schedule_created",
                        schedule=schedule,
                        is_read=False,
                    )
                )

            Notification.objects.bulk_create(notifications)

            return JsonResponse(
                {
                    "success": True,
                    "message": "Class scheduled successfully",
                    "schedule_id": schedule.id,
                    "notifications_sent": len(students),
                }
            )

        except CourseAllocation.DoesNotExist:
            return JsonResponse({"success": False, "error": "Course not found"})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"error": "Only POST allowed"})

@csrf_exempt
def get_teacher_schedules(request, teacher_id):
    session_teacher_id = request.session.get("teacher_id")
    if not session_teacher_id or int(session_teacher_id) != int(teacher_id):
        return JsonResponse({"success": False, "error": "Unauthorized"}, status=401)
    try:
        schedules = (
            ClassSchedule.objects.filter(teacher_id=teacher_id)
            .select_related("course_allocation", "course_allocation__class_assigned")
            .order_by("-date", "-start_time")
        )

        data = []
        for s in schedules:
            data.append(
                {
                    "id": s.id,
                    "class_name": s.course_allocation.class_assigned.class_name,
                    "subject": s.course_allocation.subject_name,
                    "date": s.date.strftime("%Y-%m-%d"),
                    "time": s.start_time.strftime("%H:%M"),
                    "status": s.status,
                    "status_display": s.get_status_display(),
                    "meeting_link": s.meeting_link,
                    "can_edit": s.status == "scheduled",
                    "can_delete": s.status == "scheduled",
                    "can_start": s.status == "scheduled",
                }
            )

        return JsonResponse({"success": True, "schedules": data})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})

@csrf_exempt
def update_schedule(request, schedule_id):
    if request.method == "POST":
        teacher_id = request.session.get("teacher_id")
        if not teacher_id:
            return JsonResponse({"success": False, "error": "Unauthorized"}, status=401)
        try:
            data = json.loads(request.body)
            schedule = ClassSchedule.objects.get(id=schedule_id)
            if schedule.teacher.id != teacher_id:
                return JsonResponse({"success": False, "error": "Forbidden"}, status=403)

            if schedule.status != "scheduled":
                return JsonResponse(
                    {
                        "success": False,
                        "error": "Cannot update class that is already started/completed",
                    }
                )

            old_date = schedule.date
            old_time = schedule.start_time

            if data.get("date"):
                schedule.date = data["date"]
            if data.get("start_time"):
                schedule.start_time = data["start_time"]

            schedule.save()

            if old_date != schedule.date or old_time != schedule.start_time:
                class_obj = schedule.course_allocation.class_assigned
                students = Student.objects.filter(student_class=class_obj)

                notifications = []
                for student in students:
                    notifications.append(
                        Notification(
                            sent_by_id=schedule.teacher.id,
                            student=student,
                            class_target=class_obj,
                            title=f"Schedule Updated: {schedule.course_allocation.subject_name}",
                            message=f"New time: {schedule.date} at {schedule.start_time}",
                            notification_type="schedule_updated",
                            schedule=schedule,
                            is_read=False,
                        )
                    )

                Notification.objects.bulk_create(notifications)

            return JsonResponse(
                {"success": True, "message": "Schedule updated successfully"}
            )

        except ClassSchedule.DoesNotExist:
            return JsonResponse({"success": False, "error": "Schedule not found"})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

@csrf_exempt
def delete_schedule(request, schedule_id):
    if request.method == "POST":
        teacher_id = request.session.get("teacher_id")
        if not teacher_id:
            return JsonResponse({"success": False, "error": "Unauthorized"}, status=401)
        try:
            schedule = ClassSchedule.objects.get(id=schedule_id)
            if schedule.teacher.id != teacher_id:
                return JsonResponse({"success": False, "error": "Forbidden"}, status=403)

            if schedule.status != "scheduled":
                return JsonResponse(
                    {
                        "success": False,
                        "error": "Cannot delete class that is already started/completed",
                    }
                )

            class_obj = schedule.course_allocation.class_assigned
            students = Student.objects.filter(student_class=class_obj)

            notifications = []
            for student in students:
                notifications.append(
                    Notification(
                        sent_by_id=schedule.teacher.id,
                        student=student,
                        class_target=class_obj,
                        title=f"Class Cancelled: {schedule.course_allocation.subject_name}",
                        message=f"Class on {schedule.date} at {schedule.start_time} has been cancelled",
                        notification_type="schedule_cancelled",
                        is_read=False,
                    )
                )

            Notification.objects.bulk_create(notifications)

            schedule.delete()

            return JsonResponse(
                {"success": True, "message": "Schedule deleted successfully"}
            )

        except ClassSchedule.DoesNotExist:
            return JsonResponse({"success": False, "error": "Schedule not found"})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

@csrf_exempt
def update_allocation(request, allocation_id):
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)

    teacher_id = request.session.get("teacher_id")
    if not teacher_id:
        return JsonResponse({"success": False, "error": "Unauthorized"}, status=401)

    try:
        data = json.loads(request.body)
        subject_name = data.get("subject_name")
        academic_year = data.get("academic_year")

        allocation = CourseAllocation.objects.get(id=allocation_id)
        if allocation.teacher.id != teacher_id:
            return JsonResponse({"success": False, "error": "Forbidden"}, status=403)

        teacher_id = request.session.get("teacher_id")
        if allocation.teacher.id != teacher_id:
            return JsonResponse(
                {
                    "success": False,
                    "error": "Permission denied - you do not own this allocation",
                }
            )

        has_ongoing = ClassSchedule.objects.filter(
            course_allocation=allocation, status__in=["ongoing", "completed"]
        ).exists()

        if has_ongoing:
            return JsonResponse(
                {
                    "success": False,
                    "error": "Cannot edit allocation with ongoing or completed classes",
                }
            )

        allocation.subject_name = subject_name
        allocation.academic_year = academic_year
        allocation.save()

        return JsonResponse(
            {"success": True, "message": "Allocation updated successfully"}
        )

    except CourseAllocation.DoesNotExist:
        return JsonResponse(
            {"success": False, "error": "Allocation not found"}, status=404
        )
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)

@csrf_exempt
def delete_allocation(request, allocation_id):
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)

    try:
        allocation = CourseAllocation.objects.get(id=allocation_id)

        teacher_id = request.session.get("teacher_id")
        if allocation.teacher.id != teacher_id:
            return JsonResponse(
                {
                    "success": False,
                    "error": "Permission denied - you do not own this allocation",
                }
            )

        ongoing_schedules = ClassSchedule.objects.filter(
            course_allocation=allocation, status="ongoing"
        )

        if ongoing_schedules.exists():
            return JsonResponse(
                {
                    "success": False,
                    "error": "Cannot delete allocation with ongoing classes. Please end the classes first.",
                }
            )

        schedule_count = ClassSchedule.objects.filter(
            course_allocation=allocation
        ).count()

        ClassSchedule.objects.filter(course_allocation=allocation).delete()

        subject_name = allocation.subject_name

        allocation.delete()

        return JsonResponse(
            {
                "success": True,
                "message": f'Allocation "{subject_name}" deleted successfully along with {schedule_count} scheduled classes',
                "deleted_schedules": schedule_count,
            }
        )

    except CourseAllocation.DoesNotExist:
        return JsonResponse(
            {"success": False, "error": "Allocation not found"}, status=404
        )
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)

@csrf_exempt
def start_class(request, schedule_id):
    if request.method == "POST":
        teacher_id = request.session.get("teacher_id")
        if not teacher_id:
            return JsonResponse({"success": False, "error": "Unauthorized"}, status = 401)
        try:
            schedule = ClassSchedule.objects.get(id=schedule_id)

            if schedule.teacher.id != teacher_id:
                return JsonResponse({"success": False, "error": "Unauthorized"})

            student_link = f"/live-room/?schedule={schedule.id}&role=student"
            teacher_link = f"/live-room/?schedule={schedule.id}&role=teacher"

            schedule.status = "ongoing"
            schedule.meeting_link = student_link
            schedule.save()

            class_obj = schedule.course_allocation.class_assigned
            students = Student.objects.filter(student_class=class_obj)

            notifications = []
            for student in students:
                notifications.append(
                    Notification(
                        sent_by=schedule.teacher,
                        student=student,
                        class_target=class_obj,
                        title=f"Class Started: {schedule.course_allocation.subject_name}",
                        message=f"Class is now live! Click to join.",
                        notification_type="class_started",
                        schedule=schedule,
                        is_read=False,
                    )
                )

            Notification.objects.bulk_create(notifications)

            return JsonResponse(
                {
                    "success": True,
                    "message": "Class started successfully",
                    "meeting_link": teacher_link,
                }
            )

        except ClassSchedule.DoesNotExist:
            return JsonResponse({"success": False, "error": "Schedule not found"})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

def get_teacher_allocated_classes(request, teacher_id):
    session_teacher_id = request.session.get("teacher_id")
    if not session_teacher_id or int(session_teacher_id) != int(teacher_id):
        return JsonResponse({"success": False, "error": "Unauthorized"}, status=401)
    try:
        allocations = CourseAllocation.objects.filter(
            teacher_id=teacher_id, is_active=True
        ).select_related("class_assigned")

        classes = []
        seen_class_ids = set()

        for alloc in allocations:
            class_id = alloc.class_assigned.id
            if class_id not in seen_class_ids:
                seen_class_ids.add(class_id)
                classes.append(
                    {
                        "class_id": class_id,
                        "class_name": alloc.class_assigned.class_name,
                        "display_name": alloc.class_assigned.class_name,
                    }
                )

        return JsonResponse({"success": True, "classes": classes})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)

def get_teacher_notifications(request, teacher_id):
    session_teacher_id = request.session.get("teacher_id")
    if not session_teacher_id or int(session_teacher_id) != int(teacher_id):
        return JsonResponse({"success": False, "error": "Unauthorized"}, status=401)
    try:
        notifications = (
            Notification.objects.filter(sent_by_id=teacher_id)
            .select_related("class_target")
            .order_by("-created_at")[:30]
        )

        data = []
        for n in notifications:
            class_name = n.class_target.class_name if n.class_target else "All Students"

            student_count = 0
            if n.class_target:
                student_count = Student.objects.filter(
                    student_class=n.class_target
                ).count()

            data.append(
                {
                    "id": n.id,
                    "title": n.title,
                    "message": n.message,
                    "type": n.notification_type,
                    "class_name": class_name,
                    "student_count": student_count,
                    "created_at": n.created_at.strftime("%d %b %Y, %I:%M %p"),
                }
            )

        return JsonResponse({"success": True, "notifications": data})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)

@csrf_exempt
def add_course_allocation(request):
    if request.method == "POST":
        teacher_id = request.session.get("teacher_id")
        if not teacher_id:
            return JsonResponse({"success": False, "error": "Unauthorized"}, status=401)
        try:
            data = json.loads(request.body)
            class_id = data.get("class_id")
            subject_name = data.get("subject_name")
            academic_year = data.get("academic_year", "")

            if not all([teacher_id, class_id, subject_name]):
                return JsonResponse(
                    {"success": False, "error": "Class and Subject are required"},
                    status=400,
                )

            try:
                teacher = Teacher.objects.select_related("college", "department").get(
                    id=teacher_id
                )
            except Teacher.DoesNotExist:
                return JsonResponse(
                    {"success": False, "error": "Teacher not found"}, status=404
                )

            try:
                class_obj = Class.objects.get(id=class_id)
            except Class.DoesNotExist:
                return JsonResponse(
                    {"success": False, "error": "Class not found"}, status=404
                )

            existing = CourseAllocation.objects.filter(
                teacher=teacher,
                class_assigned=class_obj,
                subject_name=subject_name,
                academic_year=academic_year,
            ).first()

            if existing:
                return JsonResponse(
                    {
                        "success": False,
                        "error": "You already teach this subject in this class",
                    },
                    status=400,
                )

            allocation = CourseAllocation.objects.create(
                teacher=teacher,
                college=teacher.college,
                department=teacher.department,
                class_assigned=class_obj,
                subject_name=subject_name,
                academic_year=academic_year,
            )

            return JsonResponse(
                {
                    "success": True,
                    "message": "Course allocated successfully",
                    "allocation_id": allocation.id,
                }
            )

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=500)

    return JsonResponse({"error": "Only POST method allowed"}, status=405)

def get_teacher_allocations(request, teacher_id):
    session_teacher_id = request.session.get("teacher_id")
    if not session_teacher_id or int(session_teacher_id) != int(teacher_id):
        return JsonResponse({"success": False, "error": "Unauthorized"}, status=401)
    try:
        allocations = CourseAllocation.objects.filter(
            teacher_id=teacher_id, is_active=True
        ).select_related(
            "class_assigned",
            "class_assigned__department",
            "class_assigned__department__college",
        )

        data = []
        for alloc in allocations:
            class_obj = alloc.class_assigned
            department = class_obj.department
            college = department.college

            data.append(
                {
                    "id": alloc.id,
                    "college_name": college.college_name,
                    "department_name": department.department_name,
                    "class_name": class_obj.class_name,
                    "subject_name": alloc.subject_name,
                    "academic_year": alloc.academic_year or "-",
                    "created_at": alloc.created_at.strftime("%d-%m-%Y"),
                }
            )

        return JsonResponse({"success": True, "allocations": data})

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)

@csrf_protect
def student_signup(request):
    print("\n" + "=" * 50)
    print("STUDENT SIGNUP FUNCTION CALLED")
    print("=" * 50)

    if request.method == "POST":
        try:
            data = request.POST
            first_name = data.get("first_name", "").strip()
            last_name = data.get("last_name", "").strip()
            email = data.get("email", "").strip()
            enrollment_no = data.get("enrollment_no", "").strip()
            dob_raw = data.get("dob", "").strip()
            password = data.get("password", "")
            confirm_password = data.get("confirm_password", "")
            terms_accepted = data.get("terms")
            parent_name = data.get("parent_name", "").strip()
            parent_email = data.get("parent_email", "").strip()
            parent_mobile = data.get("parent_mobile", "").strip()
            college_id = data.get("college_id")
            department_id = data.get("department_id")
            class_id = data.get("class_id")
            college_code = data.get("college_code")
            college_name = data.get("college_name")
            department_name = data.get("department_name")

            student_class = None

            if class_id:
                student_class = Class.objects.filter(id=class_id).first()

            print("COLLEGE:", college_id)
            print("DEPARTMENT:", department_id)
            print("CLASS:", class_id)

            if not all(
                [
                    first_name,
                    last_name,
                    email,
                    department_id,
                    class_id,
                    dob_raw,
                    password,
                    confirm_password,
                ]
            ):
                print("\n❌ MISSING REQUIRED FIELDS:")
                print(f"first_name: {first_name}")
                print(f"last_name: {last_name}")
                print(f"email: {email}")
                print(f"department_id: {department_id}")
                print(f"class_id: {class_id}")
                print(f"dob_raw: {dob_raw}")
                print(f"password: {password}")
                return JsonResponse({"error": "Missing required fields"}, status=400)

            if password != confirm_password:
                print("\n❌ PASSWORDS DO NOT MATCH")
                return JsonResponse({"error": "Passwords do not match"}, status=400)

            if Student.objects.filter(email=email).exists():
                print(f"\n❌ EMAIL ALREADY REGISTERED: {email}")
                return JsonResponse({"error": "Email already registered"}, status=400)

            if not terms_accepted:
                print("\n❌ TERMS NOT ACCEPTED")
                return JsonResponse({"error": "Terms not accepted"}, status=400)

            verified_email = request.session.get("student_email_verified")
            otp_verified_at = request.session.get("student_otp_verified_at")

            if not verified_email or verified_email != email:
                print(
                    f"\n❌ EMAIL NOT VERIFIED: Session email={verified_email}, Submitted email={email}"
                )
                return JsonResponse(
                    {"error": "Email not verified. Please complete OTP verification."},
                    status=403,
                )

            if not otp_verified_at:
                print("\n❌ OTP VERIFICATION TIME MISSING")
                return JsonResponse({"error": "OTP verification required"}, status=403)

            verified_time = datetime.fromisoformat(otp_verified_at)
            expiry_time = verified_time + timedelta(minutes=5)

            if timezone.now() > expiry_time:
                print("\n❌ OTP EXPIRED")
                request.session.pop("student_email_verified", None)
                request.session.pop("student_otp_verified_at", None)
                return JsonResponse(
                    {"error": "OTP verification expired. Please verify again."},
                    status=403,
                )

            try:
                dob = datetime.strptime(dob_raw, "%Y-%m-%d").date()
                print(f"\n✅ DOB parsed: {dob}")
            except ValueError:
                try:
                    dob = datetime.strptime(dob_raw, "%d/%m/%Y").date()
                    print(f"\n✅ DOB parsed (alternate format): {dob}")
                except ValueError:
                    print(f"\n❌ INVALID DOB FORMAT: {dob_raw}")
                    return JsonResponse({"error": "Invalid DOB"}, status=400)

            print("\n----- DATABASE CHECKS -----")

            college = None

            if college_id:
                college = College.objects.filter(id=college_id).first()
            elif college_code and college_name:
                existing_college = College.objects.filter(
                    college_name=college_name
                ).first()
                if existing_college and existing_college.college_code != college_code:
                    return JsonResponse(
                        {"error": "College name already exists with different code"},
                        status=400,
                    )
                college, created = College.objects.get_or_create(
                    college_code=college_code, defaults={"college_name": college_name}
                )

            department = None

            if department_id:
                department = Department.objects.filter(id=department_id).first()
            elif department_name and college:
                department, created = Department.objects.get_or_create(
                    college=college, department_name=department_name
                )
                print("Department created:", created)

            print("---------------------------\n")

            with transaction.atomic():
                username_part = email.split("@")[0]
                username = f"{username_part}_{int(timezone.now().timestamp())}"

                student = Student.objects.create(
                    username=username,
                    email=email,
                    password=make_password(password),
                    first_name=first_name,
                    last_name=last_name,
                    enrollment_no=enrollment_no or None,
                    college=college,
                    department=department,
                    dob=dob,
                    student_class=student_class,
                    parent_name=parent_name or None,
                    parent_email=parent_email or None,
                    parent_mobile=parent_mobile or None,
                    email_verified=True,
                    face_registered=False,
                    terms_accepted=True,
                )
                print(f"✅ Student created with ID: {student.id}")

                student.student_id = f"{username_part}{student.id}"
                student.save()
                print(f"✅ Student ID assigned: {student.student_id}")

                request.session["face_student_id"] = student.student_id
                request.session.modified = True
                request.session.save()
                print(f"✅ Session set: face_student_id = {student.student_id}")

            request.session.pop("student_email_verified", None)
            request.session.pop("student_otp_verified_at", None)

            print("\n✅ STUDENT REGISTRATION SUCCESSFUL")
            print(f"   Student ID: {student.student_id}")
            print("=" * 50 + "\n")

            return JsonResponse(
                {
                    "success": True,
                    "student_id": student.student_id,
                    "redirect": "/face-capture/",
                }
            )

        except Exception as e:
            print("\n❌ STUDENT SIGNUP ERROR:")
            print(str(e))
            print("-" * 50)
            import traceback

            traceback.print_exc()
            print("=" * 50 + "\n")
            return JsonResponse({"error": str(e)}, status=500)

    return render(request, "student/student_signup.html")

def student_login(request):
    if request.method == "POST":
        email = request.POST.get("email")
        password = request.POST.get("password")

        if not email or not password:
            return render(
                request,
                "student/student_login.html",
                {"error": "Email and password are required"},
            )

        student = Student.objects.filter(email=email).first()
        if not student:
            return render(
                request, "student/student_login.html", {"error": "Student not found"}
            )

        if not check_password(password, student.password):
            return render(
                request, "student/student_login.html", {"error": "Invalid credentials"}
            )

        if not student.is_active:
            return render(
                request, "student/student_login.html", {"error": "Account is inactive"}
            )

        request.session.flush()

        request.session["role"] = "student"
        request.session["student_id"] = student.id
        request.session["student_email"] = student.email

        return redirect("/student-dashboard/")

    return render(request, "student/student_login.html")

@student_required
def student_dashboard(request):
    print("=" * 50)
    print("STUDENT DASHBOARD VIEW CALLED")
    print("=" * 50)

    if request.session.get("role") != "student":
        print("❌ Not a student session")
        return HttpResponseRedirect("/student-login/")

    student_obj = None
    student_id = request.session.get("student_id")
    print(f"Session Student ID: {student_id}")

    if student_id:
        try:
            student_obj = Student.objects.select_related(
                "college", "department", "student_class"
            ).get(id=student_id)
            print(f"✅ Student found: {student_obj.first_name} {student_obj.last_name}")
            print(f"   College: {student_obj.college}")
            print(f"   Department: {student_obj.department}")
            print(f"   Class: {student_obj.student_class}")

            print("\n" + "=" * 50)
            print("DATABASE DEBUG INFO")
            print("=" * 50)

            total_schedules = ClassSchedule.objects.count()
            print(f"1. Total schedules in database: {total_schedules}")

            if student_obj.student_class:
                class_schedules = ClassSchedule.objects.filter(
                    course_allocation__class_assigned=student_obj.student_class
                )
                print(
                    f"2. Schedules for class '{student_obj.student_class}': {class_schedules.count()}"
                )

                for i, s in enumerate(class_schedules):
                    print(f"   Schedule {i+1}:")
                    print(f"     - Subject: {s.course_allocation.subject_name}")
                    print(f"     - Date: {s.date}")
                    print(f"     - Time: {s.start_time}")
                    print(f"     - Status: {s.status}")
                    print(f"     - Teacher: {s.teacher}")

            if student_obj.student_class:
                allocations = CourseAllocation.objects.filter(
                    class_assigned=student_obj.student_class, is_active=True
                )
                print(
                    f"\n3. Active Course Allocations for class: {allocations.count()}"
                )
                for a in allocations:
                    print(f"   - {a.subject_name} taught by {a.teacher}")

            print("=" * 50 + "\n")

        except Student.DoesNotExist:
            student_obj = None
            print("❌ Student not found in database")

    if not student_obj:
        print("❌ No student object, redirecting to login")
        return HttpResponseRedirect("/student-login/")

    today = timezone.now().date()
    print(f"Today's date: {today}")

    schedules = (
        ClassSchedule.objects.filter(
            course_allocation__class_assigned=student_obj.student_class
        )
        .select_related("teacher", "course_allocation")
        .order_by("date", "start_time")
    )

    print(f"Total schedules found with full filter: {schedules.count()}")
    for s in schedules:
        print(
            f"  - {s.course_allocation.subject_name} on {s.date} at {s.start_time} (Status: {s.status})"
        )

    today_schedules = schedules.filter(date=today)
    print(f"Today's schedules: {today_schedules.count()}")

    subjects = CourseAllocation.objects.filter(
        college=student_obj.college,
        department=student_obj.department,
        class_assigned=student_obj.student_class,
        is_active=True,
    ).select_related("teacher")

    for subject in subjects:
        print(f"Subject: {subject.subject_name}")
        print(f"Has teacher: {subject.teacher is not None}")
        if subject.teacher:
            print(f"Teacher name: {subject.teacher.full_name()}")

    print(f"Subjects found: {subjects.count()}")

    unread_count = Notification.objects.filter(
        student=student_obj, is_read=False
    ).count()

    print(f"Unread notifications: {unread_count}")

    recent_notifications = (
        Notification.objects.filter(student=student_obj)
        .select_related("sent_by", "schedule")
        .order_by("-created_at")[:5]
    )

    context = {
        "student": student_obj,
        "student_name": f"{student_obj.first_name} {student_obj.last_name}",
        "student_email": student_obj.email,
        "enrollment_no": student_obj.enrollment_no,
        "student_id": student_obj.student_id,
        "department": (
            student_obj.department.department_name if student_obj.department else ""
        ),
        "college_name": student_obj.college.college_name if student_obj.college else "",
        "class_name": (
            student_obj.student_class.class_name if student_obj.student_class else ""
        ),
        "schedules": today_schedules,
        "all_schedules": schedules,
        "subjects": subjects,
        "unread_count": unread_count,
        "recent_notifications": recent_notifications,
        "current_time": timezone.now().time().strftime("%H:%M"),
        "today_date": today.strftime("%Y-%m-%d"),
    }

    print("=" * 50)
    return render(request, "student/student_dashboard.html", context)

@csrf_protect
def student_reset_password(request):
    print("➡️ student_reset_password VIEW CALLED")

    if request.method == "POST":
        print("✅ POST request received")

        try:
            data = json.loads(request.body)
            print("📦 Request JSON:", data)

            action = data.get("action")
            email = data.get("email", "").strip()

            print("🔧 Action:", action)
            print("📧 Email:", email)

            if action == "send_otp":
                print("📨 STEP 1 : SEND OTP")

                if not email:
                    print("❌ Email missing")
                    return JsonResponse({"error": "Email required"}, status=400)

                student = Student.objects.filter(email=email).first()
                if not student:
                    print("❌ Email not registered:", email)
                    return JsonResponse({"error": "Email not registered"}, status=404)

                otp = str(random.randint(100000, 999999))
                expiry = timezone.now() + timedelta(minutes=5)

                request.session["reset_otp"] = otp
                request.session["reset_otp_email"] = email
                request.session["reset_otp_expiry"] = expiry.isoformat()
                request.session["reset_otp_role"] = "student"

                print("🔐 OTP Generated:", otp)
                print("⏰ OTP Expiry:", expiry)

                email_status = send_email_otp(
                    email, otp, purpose="forgot", role="student"
                )
                print("📤 Email send status:", email_status)

                if not email_status:
                    print("❌ Failed to send OTP email")
                    return JsonResponse({"error": "Failed to send OTP"}, status=500)

                return JsonResponse(
                    {"success": True, "message": "OTP sent successfully"}
                )

            elif action == "verify_otp":
                print("🔍 STEP 2 : VERIFY OTP")

                otp_input = data.get("otp", "").strip()

                session_otp = request.session.get("reset_otp")
                session_email = request.session.get("reset_otp_email")
                session_expiry = request.session.get("reset_otp_expiry")

                print("🔑 OTP Entered:", otp_input)
                print("📂 Session OTP:", session_otp)
                print("📧 Session Email:", session_email)
                print("⏳ Session Expiry:", session_expiry)

                if not (session_otp and session_email and session_expiry):
                    print("❌ OTP session missing")
                    return JsonResponse(
                        {"error": "OTP expired or not sent"}, status=403
                    )

                if email != session_email:
                    print("❌ Email mismatch")
                    return JsonResponse({"error": "Email mismatch"}, status=403)

                if timezone.now() > datetime.fromisoformat(session_expiry):
                    print("❌ OTP expired by time")
                    request.session.pop("reset_otp", None)
                    request.session.pop("reset_otp_email", None)
                    request.session.pop("reset_otp_expiry", None)
                    return JsonResponse({"error": "OTP expired"}, status=403)

                if otp_input != session_otp:
                    print("❌ Invalid OTP")
                    return JsonResponse({"error": "Invalid OTP"}, status=403)

                request.session["reset_email_verified"] = email
                request.session.pop("reset_otp", None)
                request.session.pop("reset_otp_expiry", None)

                print("✅ OTP verified successfully")

                return JsonResponse({"success": True, "message": "OTP verified"})

            elif action == "reset_password":
                print("🔁 STEP 3 : RESET PASSWORD")

                verified_email = request.session.get("reset_email_verified")
                print("📂 Verified email in session:", verified_email)

                if verified_email != email:
                    print("❌ OTP verification required")
                    return JsonResponse(
                        {"error": "OTP verification required"}, status=403
                    )

                new_password = data.get("password", "").strip()
                print("🔑 New password length:", len(new_password))

                if not new_password or len(new_password) < 8:
                    print("❌ Password validation failed")
                    return JsonResponse(
                        {"error": "Password must be at least 8 characters"}, status=400
                    )

                student = Student.objects.filter(email=email).first()
                if not student:
                    print("❌ Student not found")
                    return JsonResponse({"error": "Email not found"}, status=404)

                student.password = make_password(new_password)
                student.save(update_fields=["password"])

                request.session.pop("reset_email_verified", None)

                print("✅ Password reset successful for:", email)

                return JsonResponse(
                    {"success": True, "message": "Password reset successful"}
                )

            else:
                print("❌ Invalid action")
                return JsonResponse({"error": "Invalid action"}, status=400)

        except Exception as e:
            print("🔥 STUDENT RESET PASSWORD ERROR 🔥")
            print("Exception:", str(e))
            return JsonResponse({"error": "Internal server error"}, status=500)

    print("📄 Rendering student_reset_password.html")
    return render(request, "student/student_reset_password.html")

@csrf_exempt
def check_student_exists(request):
    if request.method == "POST":
        email = request.POST.get("email", "").strip()

        if email and Student.objects.filter(email=email).exists():
            return JsonResponse({"exists": True})
        else:
            return JsonResponse({"exists": False})

    return JsonResponse({"error": "Invalid request method"}, status=400)

@csrf_exempt
def get_student_schedules(request, student_id):
    session_student_id = request.session.get("student_id")
    if not session_student_id or int(session_student_id) != int(student_id):
        return JsonResponse({"success": False, "error": "Unauthorized"}, status=401)
    try:
        student = Student.objects.select_related(
            "college", "department", "student_class"
        ).get(id=student_id)

        schedules = (
            ClassSchedule.objects.filter(
                course_allocation__class_assigned=student.student_class,
            )
            .select_related("teacher", "course_allocation")
            .order_by("date", "start_time")
        )

        print(f"Total schedules found: {schedules.count()}")

        today = timezone.now().date()
        current_time = timezone.now().time()
        print(f"Today: {today}, Current Time: {current_time}")

        data = []
        for s in schedules:
            schedule_data = {
                "id": s.id,
                "subject": s.course_allocation.subject_name,
                "teacher": s.teacher.full_name() if s.teacher else "Not Assigned",
                "date": s.date.strftime("%d %b %Y"),
                "date_raw": s.date.strftime("%Y-%m-%d"),
                "time": s.start_time.strftime("%I:%M %p"),
                "time_raw": s.start_time.strftime("%H:%M"),
                "status": s.status,
                "status_display": s.get_status_display(),
                "can_join": s.status == "ongoing",
                "meeting_link": s.meeting_link if s.status == "ongoing" else None,
                "is_today": s.date == today,
                "is_upcoming": s.date > today
                or (s.date == today and s.start_time > current_time),
            }
            data.append(schedule_data)
            print(
                f"  - {schedule_data['subject']} | Date: {schedule_data['date']} | Status: {schedule_data['status']}"
            )

        return JsonResponse({"success": True, "schedules": data})

    except Student.DoesNotExist:
        print(f"❌ Student not found with ID: {student_id}")
        return JsonResponse(
            {"success": False, "error": "Student not found"}, status=404
        )
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback

        traceback.print_exc()
        return JsonResponse({"success": False, "error": str(e)}, status=500)

@csrf_exempt
def get_student_notifications(request, student_id):
    session_student_id = request.session.get("student_id")
    if not session_student_id or int(session_student_id) != int(student_id):
        return JsonResponse({"success": False, "error": "Unauthorized"}, status=401)
    try:
        student = Student.objects.get(id=student_id)

        notifications = (
            Notification.objects.filter(student=student)
            .select_related("sent_by", "schedule")
            .order_by("-created_at")
        )

        data = []
        unread_count = 0

        for n in notifications:
            if not n.is_read:
                unread_count += 1

            data.append(
                {
                    "id": n.id,
                    "title": n.title,
                    "message": n.message,
                    "type": n.notification_type,
                    "is_read": n.is_read,
                    "teacher": n.sent_by.full_name() if n.sent_by else "System",
                    "created_at": n.created_at.strftime("%d %b %Y, %I:%M %p"),
                    "schedule_id": n.schedule.id if n.schedule else None,
                    "class_name": n.class_target.class_name if n.class_target else "",
                }
            )

        print(f"Total notifications: {len(data)}, Unread: {unread_count}")

        return JsonResponse(
            {"success": True, "notifications": data, "unread_count": unread_count}
        )

    except Student.DoesNotExist:
        print(f"❌ Student not found with ID: {student_id}")
        return JsonResponse(
            {"success": False, "error": "Student not found"}, status=404
        )
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        return JsonResponse({"success": False, "error": str(e)}, status=500)

@csrf_exempt
def mark_notification_read(request, notification_id):
    # ── Only allow logged-in students to mark their own notifications ──
    session_student_id = request.session.get("student_id")
    if not session_student_id:
        return JsonResponse({"success": False, "error": "Unauthorized"}, status=401)

    if request.method == "POST":
        try:
            notification = Notification.objects.get(
                id=notification_id, student_id=session_student_id
            )
            notification.is_read = True
            notification.save()

            return JsonResponse(
                {"success": True, "message": "Notification marked as read"}
            )
        except Notification.DoesNotExist:
            print(f"❌ Notification not found: {notification_id}")
            return JsonResponse(
                {"success": False, "error": "Notification not found"}, status=404
            )
        except Exception as e:
            print(f"❌ Error: {str(e)}")
            return JsonResponse({"success": False, "error": str(e)}, status=500)

    return JsonResponse({"error": "Only POST allowed"}, status=405)

@csrf_exempt
def mark_all_notifications_read(request, student_id):
    session_student_id = request.session.get("student_id")
    if not session_student_id or int(session_student_id) != int(student_id):
        return JsonResponse({"success": False, "error": "Unauthorized"}, status=401)

    if request.method == "POST":
        try:
            count = Notification.objects.filter(
                student_id=student_id, is_read=False
            ).update(is_read=True)

            return JsonResponse(
                {"success": True, "message": f"{count} notifications marked as read"}
            )
        except Exception as e:
            print(f"❌ Error: {str(e)}")
            return JsonResponse({"success": False, "error": str(e)}, status=500)

    return JsonResponse({"error": "Only POST allowed"}, status=405)

@csrf_exempt
def check_class_access(request, schedule_id, student_id):
    # ── Session verify: URL student_id must match logged-in student ──
    session_student_id = request.session.get("student_id")
    if not session_student_id or int(session_student_id) != int(student_id):
        return JsonResponse({"success": False, "error": "Unauthorized"}, status=401)

    try:
        schedule = ClassSchedule.objects.get(id=schedule_id)
        student = Student.objects.get(id=student_id)

        print(f"Schedule Status: {schedule.status}")
        print(f"Schedule Class ID: {schedule.course_allocation.class_assigned.id}")
        print(
            f"Student Class ID: {student.student_class.id if student.student_class else 'None'}"
        )

        if (
            schedule.course_allocation.class_assigned.id == student.student_class.id
            and schedule.status == "ongoing"
        ):

            print(
                f"✅ Student CAN join class {schedule.course_allocation.subject_name}"
            )

            return JsonResponse(
                {
                    "success": True,
                    "can_join": True,
                    "meeting_link": schedule.meeting_link,
                    "subject": schedule.course_allocation.subject_name,
                    "schedule_id": schedule.id,
                }
            )
        else:
            reason = []
            if schedule.course_allocation.class_assigned.id != student.student_class.id:
                reason.append("Wrong class")
            if schedule.status != "ongoing":
                reason.append(f"Class status is {schedule.status}")

            print(f"❌ Student CANNOT join: {', '.join(reason)}")

            return JsonResponse(
                {
                    "success": True,
                    "can_join": False,
                    "message": "Class not started yet or you do not have access",
                }
            )

    except ClassSchedule.DoesNotExist:
        print(f"❌ Schedule not found: {schedule_id}")
        return JsonResponse(
            {"success": False, "error": "Schedule not found"}, status=404
        )
    except Student.DoesNotExist:
        print(f"❌ Student not found: {student_id}")
        return JsonResponse(
            {"success": False, "error": "Student not found"}, status=404
        )
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        return JsonResponse({"success": False, "error": str(e)}, status=500)

@csrf_exempt
def get_student_attendance(request, student_id):
    # ── Session verify ──
    session_student_id = request.session.get("student_id")
    if not session_student_id or int(session_student_id) != int(student_id):
        return JsonResponse({"success": False, "error": "Unauthorized"}, status=401)

    try:
        student = Student.objects.select_related(
            "college", "department", "student_class"
        ).get(id=student_id)

        enrollment_no = (student.enrollment_no or "").strip()

        # ── Get all subjects allocated to this student's class ──
        allocations = CourseAllocation.objects.filter(
            college=student.college,
            department=student.department,
            class_assigned=student.student_class,
            is_active=True,
        )

        subject_stats = []
        recent_records = []

        for allocation in allocations:
            subject = allocation.subject_name
            class_obj = allocation.class_assigned
            full_class_name = f"{class_obj.class_name} - {class_obj.department.department_name}"
            subject_class = f"{subject}({full_class_name})"
            subject_dir = os.path.join(ATTENDANCE_DIR, subject_class)

            total_classes = 0
            attended = 0
            subject_recent = []

            if os.path.exists(subject_dir) and enrollment_no:
                csv_files = sorted([
                    f for f in os.listdir(subject_dir) if f.endswith(".csv")
                ])
                for fname in csv_files:
                    fpath = os.path.join(subject_dir, fname)
                    session_counted = False
                    try:
                        with open(fpath, newline="", encoding="utf-8") as f:
                            reader = csv.DictReader(f)
                            for row in reader:
                                if not session_counted:
                                    total_classes += 1
                                    session_counted = True
                                if row.get("Enrollment", "").strip() == enrollment_no:
                                    attended += 1
                                    # Parse date for display
                                    raw_date = row.get("Date", "")
                                    try:
                                        d = datetime.strptime(raw_date, "%d-%m-%Y")
                                        display_date = d.strftime("%d %b %Y")
                                    except Exception:
                                        display_date = raw_date
                                    subject_recent.append({
                                        "date": display_date,
                                        "subject": subject,
                                        "time": row.get("Time", ""),
                                        "status": "Present",
                                    })
                    except Exception:
                        continue

            percentage = round((attended / total_classes) * 100) if total_classes > 0 else 0
            subject_stats.append({
                "name": subject,
                "total": total_classes,
                "attended": attended,
                "percentage": percentage,
            })
            recent_records.extend(subject_recent)

        # ── Sort recent by date descending, take latest 10 ──
        recent_records.sort(key=lambda x: x["date"], reverse=True)
        recent_records = recent_records[:10]

        return JsonResponse({
            "success": True,
            "attendance": {
                "subjects": subject_stats,
                "recent": recent_records,
            }
        })

    except Student.DoesNotExist:
        return JsonResponse({"success": False, "error": "Student not found"}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)

def check_college(request):
    code = request.GET.get("college_code")

    try:
        college = College.objects.get(college_code=code)

        return JsonResponse(
            {
                "exists": True,
                "college_id": college.id,
                "college_name": college.college_name,
            }
        )

    except College.DoesNotExist:
        return JsonResponse({"exists": False})

def get_departments(request, college_id):
    departments = Department.objects.filter(college_id=college_id)

    data = []

    for d in departments:
        data.append({"id": d.id, "name": d.department_name})

    return JsonResponse({"departments": data})

def get_classes(request, department_id):
    classes = Class.objects.filter(department_id=department_id)

    data = []

    for c in classes:
        data.append({"id": c.id, "name": c.class_name})

    return JsonResponse({"classes": data})

@csrf_exempt
def add_college(request):
    if request.method == "POST":

        data = json.loads(request.body)

        code = data.get("college_code")
        name = data.get("college_name")

        college, created = College.objects.get_or_create(
            college_code=code, defaults={"college_name": name}
        )

        return JsonResponse({"college_id": college.id, "created": created})

@csrf_exempt
def add_department(request):
    if request.method == "POST":

        data = json.loads(request.body)

        college_id = data.get("college_id")
        department_name = data.get("department_name")

        department, created = Department.objects.get_or_create(
            college_id=college_id, department_name=department_name
        )

        return JsonResponse({"department_id": department.id, "created": created})

@csrf_exempt
def add_class(request):
    if request.method == "POST":

        data = json.loads(request.body)

        department_id = data.get("department_id")
        class_name = data.get("class_name")

        cls, created = Class.objects.get_or_create(
            department_id=department_id, class_name=class_name
        )

        return JsonResponse({"class_id": cls.id, "created": created})

@csrf_exempt
def send_class_notification(request):
    if request.method == "POST":
        # ── Session auth check ──
        teacher_id = request.session.get("teacher_id")
        if not teacher_id:
            return JsonResponse({"success": False, "error": "Unauthorized"}, status=401)

        try:
            data = json.loads(request.body)
            class_id = data.get("class_id")
            title = data.get("title")
            message = data.get("message")
            notification_type = data.get("notification_type", "teacher_notification")

            if not all([teacher_id, class_id, title, message]):
                print("❌ Missing fields!")
                return JsonResponse(
                    {"success": False, "error": "All fields are required"}, status=400
                )

            try:
                teacher = Teacher.objects.get(id=teacher_id)
                print(f"✅ Teacher found: {teacher.full_name} (ID: {teacher.id})")
            except Teacher.DoesNotExist:
                print(f"❌ Teacher not found with ID: {teacher_id}")
                return JsonResponse(
                    {"success": False, "error": "Teacher not found"}, status=404
                )

            try:
                class_obj = Class.objects.get(id=class_id)
                print(f"✅ Class found: {class_obj.class_name} (ID: {class_obj.id})")
            except Class.DoesNotExist:
                print(f"❌ Class not found with ID: {class_id}")
                return JsonResponse(
                    {"success": False, "error": "Class not found"}, status=404
                )

            students = Student.objects.filter(student_class=class_obj)
            print(f"👥 Students in class: {students.count()}")

            if students.count() == 0:
                print("⚠️ No students found in this class!")
                return JsonResponse(
                    {
                        "success": True,
                        "message": "No students in this class",
                        "count": 0,
                    }
                )

            notifications = []
            for student in students:
                print(
                    f"  - Creating notification for student: {student.first_name} (ID: {student.id})"
                )
                notifications.append(
                    Notification(
                        sent_by=teacher,
                        student=student,
                        class_target=class_obj,
                        title=title,
                        message=message,
                        notification_type=notification_type,
                        is_read=False,
                    )
                )

            if notifications:
                created = Notification.objects.bulk_create(notifications)
                print(f"✅ Created {len(created)} notifications in database!")
                print(f"   IDs: {[n.id for n in created[:5]]}")
            else:
                print("❌ No notifications created!")

            return JsonResponse(
                {
                    "success": True,
                    "message": f"Notification sent to {len(students)} students in {class_obj.class_name}",
                    "count": len(students),
                }
            )

        except Exception as e:
            print(f"❌❌❌ ERROR: {str(e)}")
            import traceback

            traceback.print_exc()
            return JsonResponse({"success": False, "error": str(e)}, status=500)

    return JsonResponse({"error": "Only POST allowed"}, status=405)

def live_room(request):
    schedule_id = request.GET.get("schedule", "")
    role = request.GET.get("role", "student")

    student_id = request.session.get("student_id", "")
    teacher_id = request.session.get("teacher_id", "")

    if role == "teacher" and not teacher_id:
        return redirect("/teacher-login/")
    if role == "student" and not student_id:
        return redirect("/student-login/")

    subject_name = ""
    class_name = ""
    user_name = ""

    if schedule_id:
        try:
            from notifications.models import ClassSchedule

            schedule = ClassSchedule.objects.select_related(
                "course_allocation", "course_allocation__class_assigned"
            ).get(id=schedule_id)
            subject_name = schedule.course_allocation.subject_name
            class_name = schedule.course_allocation.class_assigned.class_name
            if role == "student" and schedule.status != "ongoing":
                return redirect("/student-dashboard/")
        except Exception as e:
            print(f"Schedule fetch error: {e}")

    if role == "student" and student_id:
        try:
            from students.models import Student

            s = Student.objects.get(id=student_id)
            user_name = f"{s.first_name} {s.last_name}"
        except:
            pass
    elif role == "teacher" and teacher_id:
        try:
            from teachers.models import Teacher

            t = Teacher.objects.get(id=teacher_id)
            user_name = f"{t.first_name} {t.last_name}"
        except:
            pass

    context = {
        "schedule_id": schedule_id,
        "role": role,
        "user_id": str(student_id) if student_id else str(teacher_id),
        "user_name": user_name,
        "subject_name": subject_name,
        "class_name": class_name,
    }
    return render(request, "other/live_room.html", context)

def send_email_otp(receiver_email, otp, purpose="signup", role="student"):
    EMAIL_HOST_USER = os.getenv("EMAIL_HOST_USER")
    EMAIL_HOST_PASSWORD = os.getenv("EMAIL_HOST_PASSWORD")

    if purpose == "signup":
        subject = f"MirrorMind | {role.capitalize()} Email Verification OTP"
    else:
        subject = f"MirrorMind | {role.capitalize()} Password Reset OTP"

    if purpose == "signup" and role == "student":
        body = f"""
Hello Student 👋,

Welcome to MirrorMind 🎓

Your Email Verification OTP is:

🔐 {otp}

Please enter this OTP to complete your student registration.
This OTP is valid for 5 minutes.

⚠️ Do not share this OTP with anyone.

– MirrorMind Team
"""

    elif purpose == "signup" and role == "teacher":
        body = f"""
Hello Teacher 👨‍🏫,

Welcome to MirrorMind – Smart Classroom Platform.

Your Email Verification OTP is:

🔐 {otp}

Verify your email to activate your teacher account.
This OTP is valid for 5 minutes.

⚠️ Do not share this OTP with anyone.

– MirrorMind Team
"""

    elif purpose == "forgot" and role == "student":
        body = f"""
Hello Student 👋,

We received a request to reset your MirrorMind password.

Your Password Reset OTP is:

🔐 {otp}

Use this OTP to set a new password.
This OTP is valid for 5 minutes.

If you did not request this, please ignore this email.

– MirrorMind Team
"""

    elif purpose == "forgot" and role == "teacher":
        body = f"""
Hello Teacher 👨‍🏫,

A password reset request was initiated for your MirrorMind account.

Your Password Reset OTP is:

🔐 {otp}

Use this OTP to reset your password.
This OTP is valid for 5 minutes.

If this wasn't you, please ignore this email.

– MirrorMind Team
"""

    else:
        return False

    msg = MIMEText(body)
    msg["Subject"] = subject
    msg["From"] = EMAIL_HOST_USER
    msg["To"] = receiver_email

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(EMAIL_HOST_USER, EMAIL_HOST_PASSWORD)
            server.send_message(msg)
        return True

    except Exception as e:
        print("EMAIL OTP ERROR:", e)
        return False

@csrf_protect
def email_otp_handler(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
        except Exception:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        action = data.get("action")
        email = data.get("email")
        purpose = data.get("purpose")
        role = data.get("role", "student")

        if not action or not email or not purpose:
            return JsonResponse({"error": "Missing required fields"}, status=400)

        if action == "send_otp":

            if purpose == "signup":
                if role == "student" and Student.objects.filter(email=email).exists():
                    return JsonResponse(
                        {"error": "Email already registered"}, status=400
                    )

                if role == "teacher" and Teacher.objects.filter(email=email).exists():
                    return JsonResponse(
                        {"error": "Email already registered"}, status=400
                    )

            if purpose == "forgot":
                if (
                    role == "student"
                    and not Student.objects.filter(email=email).exists()
                ):
                    return JsonResponse({"error": "Email not registered"}, status=400)

                if (
                    role == "teacher"
                    and not Teacher.objects.filter(email=email).exists()
                ):
                    return JsonResponse({"error": "Email not registered"}, status=400)

            otp = str(random.randint(100000, 999999))
            expiry = timezone.now() + timedelta(minutes=5)

            request.session["otp_email"] = email
            request.session["otp_code"] = otp
            request.session["otp_expiry"] = expiry.isoformat()
            request.session["otp_purpose"] = purpose
            request.session["otp_role"] = role

            mail_sent = send_email_otp(email, otp, purpose, role)

            if not mail_sent:
                return JsonResponse({"error": "Failed to send OTP"}, status=500)

            return JsonResponse({"success": True})

        elif action == "verify_otp":
            otp_input = data.get("otp")

            if not otp_input:
                return JsonResponse(
                    {"verified": False, "error": "OTP required"}, status=400
                )

            # ── Brute force limit ──
            attempts = request.session.get("otp_attempts", 0)
            if attempts >= 5:
                request.session.flush()
                return JsonResponse({"verified": False, "error": "Too many attempts. Request new OTP."}, status=429)

            session_otp     = request.session.get("otp_code")
            session_email   = request.session.get("otp_email")
            session_expiry  = request.session.get("otp_expiry")
            session_purpose = request.session.get("otp_purpose")
            session_role    = request.session.get("otp_role")

            if not all([session_otp, session_email, session_expiry]):
                return JsonResponse(
                    {"verified": False, "error": "OTP expired or not sent"}, status=400
                )

            if email != session_email:
                return JsonResponse(
                    {"verified": False, "error": "Email mismatch"}, status=400
                )

            if timezone.now() > datetime.fromisoformat(session_expiry):
                request.session.flush()
                return JsonResponse(
                    {"verified": False, "error": "OTP expired"}, status=400
                )

            if otp_input != session_otp:
                request.session["otp_attempts"] = attempts + 1
                return JsonResponse(
                    {"verified": False, "error": "Invalid OTP"}, status=400
                )

            # ── Success — reset attempts ──
            request.session.pop("otp_attempts", None)

            if session_purpose == "signup":
                if session_role == "student":
                    request.session["student_email_verified"] = email
                    request.session["student_otp_verified_at"] = (
                        timezone.now().isoformat()
                    )
                else:
                    request.session["teacher_email_verified"] = email
                    request.session["teacher_otp_verified_at"] = (
                        timezone.now().isoformat()
                    )

            elif session_purpose == "forgot":
                request.session["reset_email_verified"] = email

            request.session.pop("otp_code", None)
            request.session.pop("otp_expiry", None)

            return JsonResponse({"verified": True})

    return JsonResponse({"error": "Invalid action"}, status=400)

def create_schedule_notifications(schedule, notification_type):
    try:
        class_target = schedule.course_allocation.class_assigned
        students = Student.objects.filter(student_class=class_target)

        if not students.exists():
            return True

        title_map = {
            "schedule_created": f"New Class Scheduled: {schedule.course_allocation.subject_name}",
            "schedule_updated": f"Schedule Updated: {schedule.course_allocation.subject_name}",
            "schedule_cancelled": f"Class Cancelled: {schedule.course_allocation.subject_name}",
            "class_started": f"Class Started: {schedule.course_allocation.subject_name}",
        }

        message_map = {
            "schedule_created": f"Class scheduled for {schedule.date} at {schedule.start_time}",
            "schedule_updated": f"New time: {schedule.date} at {schedule.start_time}",
            "schedule_cancelled": f"Class on {schedule.date} at {schedule.start_time} has been cancelled",
            "class_started": f"Class is now live! Click to join.",
        }

        title = title_map.get(notification_type, "Notification")
        message = message_map.get(notification_type, "")

        if schedule.meeting_link and notification_type == "class_started":
            message += f"\nJoin link: {schedule.meeting_link}"

        notifications = []
        for student in students:
            notifications.append(
                Notification(
                    sent_by=schedule.teacher,
                    student=student,
                    class_target=class_target,
                    title=title,
                    message=message,
                    notification_type=notification_type,
                    schedule=schedule,
                    is_read=False,
                )
            )

        Notification.objects.bulk_create(notifications)
        return True
    except Exception as e:
        print(f"Error creating notifications: {e}")
        return False

@csrf_exempt
def end_class(request, schedule_id):
    if request.method == "POST":
        try:
            schedule = ClassSchedule.objects.get(id=schedule_id)

            teacher_id = request.session.get("teacher_id")
            if not teacher_id or schedule.teacher.id != teacher_id:
                return JsonResponse(
                    {"success": False, "error": "Unauthorized"}, status=403
                )

            if schedule.status != "ongoing":
                return JsonResponse({"success": False, "error": "Class is not ongoing"})

            schedule.status = "completed"
            schedule.save()

            class_obj = schedule.course_allocation.class_assigned
            students = Student.objects.filter(student_class=class_obj)
            notifications = []
            for student in students:
                notifications.append(
                    Notification(
                        sent_by=schedule.teacher,
                        student=student,
                        class_target=class_obj,
                        title=f"Class Ended: {schedule.course_allocation.subject_name}",
                        message=f"Class has ended. See you next time!",
                        notification_type="class_ended",
                        schedule=schedule,
                        is_read=False,
                    )
                )
            Notification.objects.bulk_create(notifications)

            return JsonResponse(
                {"success": True, "message": "Class ended successfully"}
            )

        except ClassSchedule.DoesNotExist:
            return JsonResponse({"success": False, "error": "Schedule not found"})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"error": "Only POST allowed"}, status=405)

@csrf_exempt
def detect_emotion(request):
    if request.method != "POST":
        return JsonResponse({"error": "POST only"}, status=405)

    # ── Session auth check ──
    if not request.session.get("student_id") and not request.session.get("teacher_id"):
        return JsonResponse({"error": "Unauthorized"}, status=401)

    try:
        data = json.loads(request.body)
        image_data = data.get("image", "")
        student_id = data.get("student_id", "unknown")

        if not image_data:
            return JsonResponse({"error": "No image"}, status=400)

        if "," in image_data:
            image_data = image_data.split(",")[1]
        img_bytes = base64.b64decode(image_data)
        np_arr = np.frombuffer(img_bytes, np.uint8)
        frame = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)

        if frame is None:
            return JsonResponse({"emotion": "neutral", "confidence": 0.7})

        h, w = frame.shape[:2]
        if w > 320:
            scale = 320 / w
            frame = cv2.resize(frame, (320, int(h * scale)))

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(60, 60))

        face_img = frame
        if len(faces) > 0:
            largest = max(faces, key=lambda r: r[2] * r[3])
            x, y, w_f, h_f = largest
            pad_x = int(w_f * 0.15)
            pad_y = int(h_f * 0.15)
            x1 = max(0, x - pad_x)
            y1 = max(0, y - pad_y)
            x2 = min(frame.shape[1], x + w_f + pad_x)
            y2 = min(frame.shape[0], y + h_f + pad_y)
            face_img = frame[y1:y2, x1:x2]

        result = DeepFace.analyze(
            face_img,
            actions=["emotion"],
            enforce_detection=False,
            detector_backend="opencv",
            silent=True,
        )

        if isinstance(result, list):
            result = result[0]

        emotions = result.get("emotion", {})
        dominant = result.get("dominant_emotion", "neutral")

        EMOTION_MAP = {
            "happy":     ("Happy",   0),
            "surprised": ("Happy",   0),
            "neutral":   ("Focused", 1),
            "disgusted": ("Neutral", 1),
            "angry":     ("Bored",   2),
            "fearful":   ("Bored",   2),
            "sad":       ("Sleepy",  3),
        }

        mapped_label, _ = EMOTION_MAP.get(dominant, ("Neutral", 1))
        raw_conf = emotions.get(dominant, 70) / 100.0
        confidence = round(min(0.99, max(0.5, raw_conf)), 3)

        print(f"✅ DeepFace Emotion [{student_id}]: {dominant} → {mapped_label} ({confidence:.2f})")

        return JsonResponse({
            "emotion": dominant,
            "mapped_emotion": mapped_label,
            "confidence": confidence,
            "all_emotions": {k: round(v / 100, 3) for k, v in emotions.items()},
        })

    except Exception as e:
        print("❌ Emotion detection error:", e)
        return JsonResponse({"emotion": "neutral", "confidence": 0.7})

def term_condition(request):
    return render(request, "other/term_condition.html")

def help_contact(request):
    return render(request, "other/help_contact.html")

def about(request):
    return render(request, "other/about.html")

def privacy_policy(request):
    return render(request, "other/privacy_policy.html")

@csrf_exempt
def export_attendance_excel(request):
    teacher_id = request.session.get("teacher_id")
    if not teacher_id:
        return JsonResponse({"error": "Not logged in"}, status=401)

    subject = request.GET.get("subject", "")
    date_input = request.GET.get("date", "")

    if not subject or not date_input:
        return JsonResponse({"error": "Subject and date required"}, status=400)

    try:
        formatted_date = datetime.strptime(date_input, "%Y-%m-%d").strftime("%d-%m-%Y")
        allocation = CourseAllocation.objects.filter(
            teacher_id=teacher_id, subject_name=subject
        ).select_related("class_assigned", "class_assigned__department").first()

        if not allocation:
            return JsonResponse({"error": "Subject not found"}, status=404)

        class_obj = allocation.class_assigned
        full_class_name = f"{class_obj.class_name} - {class_obj.department.department_name}"
        subject_class = f"{subject}({full_class_name})"
        subject_dir   = os.path.join(ATTENDANCE_DIR, subject_class)

        # ── Collect all CSV files for this date (handles _S1, _S2 suffixes) ──
        if not os.path.exists(subject_dir):
            return JsonResponse({"error": "Attendance records not found"}, status=404)

        matching_files = sorted([
            f for f in os.listdir(subject_dir)
            if f.startswith(f"Attendance_{formatted_date}") and f.endswith(".csv")
        ])

        if not matching_files:
            return JsonResponse({"error": "Attendance file not found for this date"}, status=404)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")

        ws.append(["MirrorMind - Attendance Report"])
        ws.append(["Subject:", subject])
        ws.append(["Class:", full_class_name])
        ws.append(["Date:", date_input])
        ws.append([])

        headers = ["#", "Subject(Class)", "Enrollment No", "Student Name", "Date", "Time"]
        ws.append(headers)
        for i, cell in enumerate(ws[6]):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # ── Merge all session files, deduplicate by enrollment ──
        seen_enrollments = set()
        row_idx = 1
        for fname in matching_files:
            fpath = os.path.join(subject_dir, fname)
            with open(fpath, newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    enrollment = row.get("Enrollment", "").strip()
                    if enrollment and enrollment not in seen_enrollments:
                        seen_enrollments.add(enrollment)
                        ws.append([
                            row_idx,
                            row.get("Subject(Class)", ""),
                            enrollment,
                            row.get("Name", ""),
                            row.get("Date", ""),
                            row.get("Time", ""),
                        ])
                        row_idx += 1

        ws.append([])
        ws.append(["Total Present:", row_idx - 1])

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="Attendance_{subject}_{date_input}.xlsx"'
        wb.save(response)
        return response

    except Exception as e:
        print("EXCEL EXPORT ERROR:", e)
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
def analytics_data(request):
    teacher_id = request.session.get("teacher_id")
    if not teacher_id:
        return JsonResponse({"SUCCESS": False, "ERROR": "Unauthorized"}, status=401)

    subject    = request.GET.get("subject", "")

    if not teacher_id or not subject:
        return JsonResponse({"success": False, "error": "Missing params"}, status=400)

    try:
        allocation = CourseAllocation.objects.filter(
            teacher_id=teacher_id, subject_name=subject
        ).select_related("class_assigned", "class_assigned__department").first()

        if not allocation:
            return JsonResponse({"success": False, "error": "Allocation not found"})

        class_obj = allocation.class_assigned
        full_class_name = f"{class_obj.class_name} - {class_obj.department.department_name}"
        subject_class   = f"{subject}({full_class_name})"
        subject_dir     = os.path.join(ATTENDANCE_DIR, subject_class)

        trend_labels, trend_data = [], []
        total_present_all, class_count = 0, 0

        if os.path.exists(subject_dir):
            files = sorted([f for f in os.listdir(subject_dir) if f.endswith(".csv")])[-10:]
            for fname in files:
                fpath = os.path.join(subject_dir, fname)
                with open(fpath, newline="", encoding="utf-8") as f:
                    reader = csv.DictReader(f)
                    count = sum(1 for _ in reader)
                try:
                    date_str = fname.replace("Attendance_", "").replace(".csv", "")
                    # Strip session suffix like _S1, _S2 if present
                    import re as _re
                    date_str = _re.sub(r'_S[0-9]+$', '', date_str)
                    d = datetime.strptime(date_str, "%d-%m-%Y")
                    trend_labels.append(d.strftime("%d %b"))
                except Exception:
                    trend_labels.append(fname[:8])
                trend_data.append(count)
                total_present_all += count
                class_count += 1

        avg_present = round(total_present_all / class_count) if class_count else 0
        total_students = Student.objects.filter(student_class=class_obj).count()

        schedules_qs = ClassSchedule.objects.filter(
            course_allocation=allocation
        ).values_list('id', flat=True)

        emotion_logs = EmotionLog.objects.filter(schedule_id__in=schedules_qs)
        emotion_counter = Counter(log.emotion for log in emotion_logs)

        emotion_labels = ['Happy', 'Focused', 'Neutral', 'Bored', 'Sad', 'Sleepy', 'Unavailable']
        emotion_data   = [emotion_counter.get(e, 0) for e in emotion_labels]
        emotion_colors = ['#28a745', '#007bff', '#ffc107', '#fd7e14', '#6496ff', '#dc3545', '#adb5bd']

        students_in_class = Student.objects.filter(student_class=class_obj)
        per_student = []
        for stu in students_in_class:
            stu_logs = emotion_logs.filter(student=stu)
            total = stu_logs.count()
            if total == 0:
                per_student.append({
                    "name": f"{stu.first_name} {stu.last_name}",
                    "enrollment": stu.enrollment_no,
                    "dominant": "No Data",
                    "engagement_pct": 0,
                    "emotions": {}
                })
                continue

            stu_counter = Counter(l.emotion for l in stu_logs)
            dominant    = stu_counter.most_common(1)[0][0]
            engaged     = stu_counter.get('Happy', 0) + stu_counter.get('Focused', 0)
            eng_pct     = round((engaged / total) * 100)

            per_student.append({
                "name": f"{stu.first_name} {stu.last_name}",
                "enrollment": stu.enrollment_no,
                "dominant": dominant,
                "engagement_pct": eng_pct,
                "emotions": {k: round((v / total) * 100) for k, v in stu_counter.items()}
            })

        per_student.sort(key=lambda x: x['engagement_pct'])

        return JsonResponse({
            "success":         True,
            "avg_present":     avg_present,
            "total_classes":   total_students if total_students else 30,
            "trend_labels":    trend_labels,
            "trend_data":      trend_data,
            "class_count":     class_count,
            "emotion_labels":  emotion_labels,
            "emotion_data":    emotion_data,
            "emotion_colors":  emotion_colors,
            "per_student":     per_student,
        })

    except Exception as e:
        import traceback; traceback.print_exc()
        print("ANALYTICS ERROR:", e)
        return JsonResponse({"success": False, "error": str(e)}, status=500)

conversation_history = {}

@csrf_exempt
def study_chatbot(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "response": "Only POST allowed"})

    try:
        data = json.loads(request.body)
        question = data.get("question", "").strip()
        session_id = request.session.session_key or "default"

        if not question:
            return JsonResponse({"success": False, "response": "Please ask a question"})

        API_KEYS = getattr(settings, "GENAI_API_KEYS", [])

        if not API_KEYS:
            return JsonResponse({
                "success": False, 
                "response": "No API keys configured"
            })

        if session_id not in conversation_history:
            conversation_history[session_id] = []

        system_prompt = """
You are MirrorMind Study Assistant, an advanced AI helping students.

Rules:
- Answer clearly and concisely
- Use headings for organization
- Use bullet points for lists
- Explain step by step
- Provide examples where helpful
- Be friendly and encouraging
"""

        conversation_history[session_id].append(f"User: {question}")
        history_text = "\n".join(conversation_history[session_id][-6:])

        prompt = f"""
{system_prompt}

Conversation History:
{history_text}

User Question:
{question}

Answer:
"""

        print(f"🔄 Question: {question[:60]}...")

        MODEL_NAME = "gemini-2.5-flash"

        for i, key in enumerate(API_KEYS):
            try:
                print(f"🔄 Trying key {i+1}/{len(API_KEYS)}")
                
                client = genai.Client(api_key=key)
                
                response = client.models.generate_content(
                    model=MODEL_NAME,
                    contents=prompt
                )
                
                if response and response.text:
                    answer = response.text.strip()
                    conversation_history[session_id].append(f"AI: {answer[:100]}...")
                    print(f"✅ Success with key {i+1}")
                    
                    return JsonResponse({
                        "success": True, 
                        "response": answer
                    })

            except Exception as e:
                error_msg = str(e)
                print(f"❌ Key {i+1} failed: {error_msg[:100]}")
                continue

        return JsonResponse({
            "success": False,
            "response": "All API keys failed. Please try: gemini-2.5-flash, gemini-2.0-flash, or gemini-2.0-flash-lite"
        })

    except Exception as e:
        print("❌ ERROR:", str(e))
        print(traceback.format_exc())
        return JsonResponse({
            "success": False, 
            "response": f"Server error occurred"
        })